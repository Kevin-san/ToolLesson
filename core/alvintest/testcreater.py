# -*- encoding:UTF-8 -*-
'''
Created on 2020/5/3

@author: xcKev
'''
from alvinwritecreater import excelwriter,fileswriter,pdfwriter
from alvinwritecreater.pdfwriter import pdf2images
from alvinentities.pdfitems import PdfItem
from alvinentities import excelstyles
import time
import os
from openpyxl.styles import colors,Side
curPath = os.path.abspath(os.path.dirname(__file__))
rootPath = curPath[:curPath.find("CorePdfPage\\")+len("CorePdfPage\\")]
practices_dir=F"{rootPath}practices"
font_dir=F"{rootPath}core/alvinconst"

def test_csv_writer():
    csv_file1=F"{practices_dir}/filedetails.csv"
    csv_file2=F"{practices_dir}/humans.csv"
    if os.path.isfile(csv_file1):
        os.remove(csv_file1)
    if os.path.isfile(csv_file2):
        os.remove(csv_file2)
    fieldnames1=['filename','filesize','filesuffix','filerowcount']
    fieldline1=['test.csv','100bytes','csv',1]
    fieldnames2=['name','age','sex','score','class']
    fieldlines2=[['Xiaoming','22','Man','34','2-1'],['Xiaohong','72','Woman','7','1-1']]
    csvwriter=fileswriter.CsvWriter(csv_file1,_delimiter="|")
    csvwriter.write_header(fieldnames1)
    csvwriter.write_line(fieldline1)
    csvwriter2=fileswriter.CsvWriter(csv_file2,_delimiter=",")
    csvwriter2.write_header(fieldnames2)
    csvwriter2.write_lines(fieldlines2)

def test_excel_writer():
    xlsx_file1=F"{practices_dir}/test.xlsx"
    sheet_names=['test1','sheet1','value1']
    excel_writer=excelwriter.ExcelWriter(xlsx_file1,sheet_names=sheet_names)
    excel_writer.set_current_sheet('value1')
    excel_style=excelstyles.ExcelStyle()
    rows1=['hah','value','mytest']
    rows2=[1,2,4,5,39,20]
    columns1=[1230,324432,232432,345]
    columns2=['hva','cool','clearly']
    excel_writer.write_row(1, rows1, excel_style.get_style())
    excel_writer.write_row(2, rows2, excel_style.get_style())
    excel_writer.write_column(7, columns1, excel_style.get_style())
    excel_writer.write_column(9, columns2, excel_style.get_style())
    excel_writer.write_cell_with_formula(2, 8, 'F2+G2')
    excel_writer.write_hyperlink(3, 8, 'http://www.baidu.com', 'baidu')
#     %y 两位数的年份表示（00-99）
#     %Y 四位数的年份表示（000-9999）
#     %m 月份（01-12）
#     %d 月内中的一天（0-31）
#     %H 24小时制小时数（0-23）
#     %I 12小时制小时数（01-12）
#     %M 分钟数（00=59）
#     %S 秒（00-59）
#     %a 本地简化星期名称
#     %A 本地完整星期名称
#     %b 本地简化的月份名称
#     %B 本地完整的月份名称
#     %c 本地相应的日期表示和时间表示
#     %j 年内的一天（001-366）
#     %p 本地A.M.或P.M.的等价符
#     %U 一年中的星期数（00-53）星期天为星期的开始
#     %w 星期（0-6），星期天为星期的开始
#     %W 一年中的星期数（00-53）星期一为星期的开始
#     %x 本地相应的日期表示
#     %X 本地相应的时间表示
#     %Z 当前时区的名称
#     %% %号本身
    excel_writer.write_datetime(3, 9, "%b %d %Y %H:%M:%S", time.localtime(), excel_style.get_style())
    excel_writer.write_cell(3, 10, "SKT-109")
    excel_writer.set_current_sheet("test1")
    #_name='Times New Roman', _size=12, _bold=False, _strike=False, _italic=False, _color=colors.BLACK
    font=excelstyles.ExcelFont(_name='Comic Sans MS',_size=15,_bold=True,_color=colors.BLUE)
    # Side(border_style='thin', color='000000')  _left_side, _right_side, _top_side, _bottom_side
    each_side=Side(border_style='double',color='000000')
    border=excelstyles.ExcelBorders(_left_side=each_side,_right_side=each_side,_top_side=each_side, _bottom_side=each_side)
    #fill_type=None, start_color='FFFFFF', end_color='000000'
    pattern=excelstyles.ExcelPattern(fill_type='solid',start_color=colors.RED)
    align=excelstyles.ExcelAlignment(_horz='center',_vert='center')
    new_excel_style=excelstyles.ExcelStyle(font=font, border=border, protection=None, alignment=align, pattern_fill=pattern)
    excel_writer.merge_cells(1, 10, 1, 10, 100, new_excel_style.get_style())
    excel_writer.set_current_sheet("sheet1")
    excel_writer.merge_cells(1, 10, 1, 10, "Welcome Sheet1", excel_style.get_style())
    excel_writer.unmerge_cells(1, 10, 1, 10)
    excel_writer.save()

def test_pdf_writer():
    pdf_file1=F"{practices_dir}/test.pdf"
    pdf_writer=pdfwriter.SimplePdfWriter(pdf_file1)
    pdf_writer.register_font("楷体", F"{font_dir}/SIMKAI.TTF")
    image_path="I:/图片/linux/linux_install7_img_008.png"

    page_item=PdfItem.image(image_path,120,120)
    page_text=PdfItem.paragraph("Hello World!", text_style=pdf_writer.stylesheet['Normal'])
    print(type(pdf_writer.get_image(page_item)))
    stories = [[pdf_writer.get_paragraph(page_text),pdf_writer.get_image(page_item)]]
    pdf_writer.write_pages(stories)

def test_pdf2images():
    pdf2images('E:/lib/books/[精通正则表达式(第三版)].（美）佛瑞德.扫描版.pdf')

if __name__=="__main__":
    test_csv_writer()
    test_excel_writer()