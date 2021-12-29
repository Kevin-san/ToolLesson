# -*- encoding:UTF-8 -*-
'''
Created on 2020/5/3

@author: xcKev
'''
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from xlutils import copy
class ExcelReader(object):
    
    def __init__(self, filename):
        self._filename = filename
        self._workbook = openpyxl.load_workbook(self._filename, data_only=True)
        self._sheet_dict = {}
        for sheet_name in self._workbook.sheetnames:
            self._sheet_dict[sheet_name] = self._workbook.get_sheet_by_name(sheet_name)
        self._sheet = self._workbook.get_sheet_by_name(self._workbook.sheetnames[0])
        self._rows = self._sheet.rows
        self._cols = self._sheet.columns
        
    def set_current_sheet(self, sheet_name):
        self._sheet = self._sheet_dict[sheet_name]
        self._rows = list(self._sheet.rows)
        self._cols = list(self._sheet.columns)
        
    def get_row_cnt(self):
        return len(self._rows)
    
    def get_col_cnt(self):
        return len(self._cols)
    
    def get_rows(self, row_id):
        return self._rows[row_id]
    
    def get_rows_value(self, row_id):
        return self._rows[row_id].value
    
    def get_cols(self, col_id):
        return self._cols(col_id)
    
    def get_cols_value(self, col_id):
        return self._cols(col_id).value
    
    def get_cell_by_id(self, row_id, col_id):
        return self._sheet.cell(row=row_id, column=col_id)
    
    def get_cell_by_str(self, cell_str):
        return self._sheet[cell_str]
    
    def get_cells_by_str(self, begin_str, end_str):
        return self._sheet[begin_str:end_str]
    
    def get_col_str_by_id(self, col_id):
        return get_column_letter(col_id)
    
    def get_col_id_by_str(self, col_str):
        return column_index_from_string(col_str)
    
    def get_hyperlink(self, row_id, col_id):
        link = self._sheet.hyperlink_map.get((row_id, col_id))
        url = '(No URL)'
        if link:
            url = link.url_or_path
        return url
    
    def copy(self, new_file_name):
        wb = copy(self._workbook)
        wb.save(new_file_name)
