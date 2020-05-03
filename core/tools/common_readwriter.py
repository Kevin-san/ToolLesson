#-*- encoding:UTF-8 -*-
'''
Created on 2019/6/1

@author: xcKev
'''
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Protection, Alignment, Border, Side, NamedStyle, colors
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.hyperlink import Hyperlink
from xlutils import copy
from PyPDF2.pdf import PdfFileReader, PdfFileWriter
from reportlab.lib.units import inch,mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Table,Paragraph,SimpleDocTemplate,Spacer,Image,TableStyle,PageBreak
from reportlab.lib.pagesizes import letter,A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.lineplots import LinePlot
from reportlab.graphics.charts.textlabels import Label
from reportlab.graphics.shapes import Drawing,Rect,Circle
from reportlab.lib.colors import HexColor
from reportlab.graphics.widgets.markers import makeMarker
from reportlab.graphics.charts.legends import Legend


def get_csv_columns(in_file,seperator_char,from_row_id,to_row_id,column_id):
    in_file_h = open(in_file,'rt')
    csv.register_dialect('mydialect',delimiter=seperator_char,quoting=csv.QUOTE_MINIMAL)
    csv_read_h=csv.reader(in_file_h,'mydialect')
    org_lines=[]
    for line in csv_read_h:
        org_lines.append(line)
    result_lines=[]
    from_row_id=int(from_row_id)
    to_row_id=int(to_row_id)
    column_id=int(column_id)
    if to_row_id == -1:
        to_row_id=int(len(org_lines))
    for row_id in range(from_row_id,to_row_id):
        result_line=org_lines[row_id]
        result_lines.append(result_line[column_id])
    return result_lines


class CsvReader(object):
    
    def __init__(self,filename,_delimiter=',',_quoting=csv.QUOTE_ALL):
        self._filename=filename
        self.set_csv_format(_delimiter, _quoting)
        self._file_reader=open(self._filename,'rt')
        self._lines=self.read_lines()
        self._total_cnt=self.get_total_line_num()
    
    def get_csv_reader(self):
        return csv.reader(self._file_reader,'reader_dialect')
    
    def set_csv_format(self,_delimiter,_quoting):
        csv.register_dialect('reader_dialect',delimiter=_delimiter,quoting=_quoting)
        
    def read_lines(self):
        lines=[]
        reader=self.get_csv_reader()
        for line in reader:
            lines.append(line)
        return lines
    def get_total_line_num(self):
        if self._lines:
            return len(self._lines)
        return 0
    
    def get_headers(self,index=0):
        if self._lines:
            return self._lines[index]
        return []
    
    def get_details(self,begin_index,end_index):
        details=[]
        if begin_index < end_index and end_index <= self._total_cnt:
            for i in range(begin_index,end_index):
                details.append(self._lines[i])
        return details

class CsvWriter(object):
    def __init__(self,filename,_delimiter=',',_quoting=csv.QUOTE_ALL):
        self._filename=filename
        self.set_csv_format(_delimiter, _quoting)
        self._file_writer=open(self._filename,'a+',newline='')
        
    def get_csv_writer(self):
        return csv.writer(self._file_writer,'writer_dialect')
    
    def set_csv_format(self,_delimiter,_quoting):
        csv.register_dialect('writer_dialect',delimiter=_delimiter,quoting=_quoting)
        
    def write_header(self,fieldnames):
        writer=csv.DictWriter(self._file_writer,dialect='writer_dialect',fieldnames=fieldnames)
        writer.writeheader()
    
    def write_line(self,line):
        csv_writer=self.get_csv_writer()
        csv_writer.writerow(line)
    
    def write_lines(self,lines):
        csv_writer=self.get_csv_writer()
        csv_writer.writerows(lines)
        
class ExcelFont(object):
    
    def __init__(self,_name='Times New Roman',_size=12,_bold=False,_strike=False,_italic=False,_color=colors.BLACK):
        self._font=Font()
        self._font.name=_name
        self._font.size=_size
        self._font.bold=_bold
        self._font.strike=_strike
        self._font.italic=_italic
        self._font.color=_color
        
    def get_font(self):
        return self._font
    
class ExcelBorders(object):
    
    def __init__(self,_left_side,_right_side,_top_side,_bottom_side):
        self._border = Border(left=Side(border_style='thin',color='000000'),right=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'))# Create Borders
        #'dashDot','dashDotDot','dashed','dotted','double','hair','medium','mediumDashDot','mediumDashDotDot','mediumDashed','slantDashDot','thick','thin'
        if _left_side:
            self._border.left=_left_side
        if _right_side:
            self._border.right=_right_side
        if _top_side:
            self._border.top=_top_side
        if _bottom_side:
            self._border.bottom=_bottom_side
        
    def get_borders(self):
        return self.borders
    
class ExcelPattern(object):
    def __init__(self,fill_type=None,start_color='FFFFFF',end_color='000000'):
        self.pattern = PatternFill(fill_type=fill_type,start_color=start_color,end_color=end_color)
        #fill_type# 'none'、'solid'、'darkDown'、'darkGray'、'darkGrid'、'darkHorizontal'、'darkTrellis'、'darkUp'、'darkVertical'、'gray0625'、'gray125'、'lightDown'、'lightGray'、'lightGrid'、'lightHorizontal'、'lightTrellis'、'lightUp'、'lightVertical'、'mediumGray'
        
    def get_pattern(self):
        return self.pattern
    
class ExcelAlignment(object):
    def __init__(self,_horz='left',_vert='center',_wrap_text=True):
        self.alignment = Alignment()
        self.alignment.horizontal = _horz#left,right,center,distributed,centerContinuous,justify,fill,general
        self.alignment.vertical = _vert#center,top,bottom,justify,distributed
        self.alignment.wrap_text=_wrap_text
        
    def get_alignment(self):
        return self.alignment

class ExcelProtection(object):
    
    def __init__(self,_locked=False,_hidden=False):
        self._protection=Protection(locked=_locked,hidden=_hidden)
    
    def get_protection(self):
        return self._protection
    
class ExcelStyle(object):
    
    def __init__(self,font=None,border=None,protection=None,alignment=None,pattern_fill=None):
        self._style=NamedStyle()
        if font:
            self._style.font=font
        if border:
            self._style.border=border
        if protection:
            self._style.protection=protection
        if alignment:
            self._style.alignment=alignment
        if pattern_fill:
            self._style.fill=pattern_fill
    
    def get_style(self):
        return self._style
    

class ExcelWriter(object):
     
    def __init__(self,filename,sheet_names):
        self._filename=filename
        self._workbook=openpyxl.Workbook()
        self._sheet_dict={}
        for i,sheet_name in enumerate(sheet_names):
            sheet=self._workbook.create_sheet(sheet_name, i)
            self._sheet_dict[sheet_name]=sheet
        self._sheet=self._sheet_dict[sheet_names[0]]
    
    def set_current_sheet(self,sheet_name):
        self._sheet=self.get_sheet(sheet_name)
        
    def set_style(self,cell,excel_style):
        cell.font=excel_style.font
        cell.border=excel_style.border
        cell.alignment=excel_style.alignment
        cell.fill=excel_style.fill
        cell.protection=excel_style.protection
        
    def get_sheet(self,sheet_name):
        return self._sheet_dict[sheet_name]
    
    def remove_sheet(self,sheet_name):
        return self._workbook.remove_sheet(self._sheet_dict[sheet_name])
    
    def get_cell_str(self,row_id,col_id):
        col_str=get_column_letter(col_id)
        cell_str=col_str+str(row_id)
        return cell_str
    
    def write_row(self,row_id,rows,excel_style):
        for i in range(1,len(rows)+1):
            cell_str=self.get_cell_str(row_id, i)
            self._sheet[cell_str]=rows[i-1]
            self.set_style(self._sheet[cell_str], excel_style)

    def write_column(self,column_id,columns,excel_style):
        for i in range(1,len(columns)+1):
            cell_str=self.get_cell_str(i, column_id)
            self._sheet[cell_str]=columns[i-1]
            self.set_style(self._sheet[cell_str], excel_style)

    def append_rows(self,rows):
        self._sheet.append(rows)
    
    def append_cols(self,cols):
        self._sheet.append(list(zip(cols)))
    
    def write_cell(self,row_id,column_id,cell_value):
        cell_str=self.get_cell_str(row_id, column_id)
        self._sheet[cell_str]=cell_value
    
    def write_hyperlink(self,row_id,column_id,hyperlink,hyperlink_value):
        cell_str=self.get_cell_str(row_id, column_id)
        hyper_link=Hyperlink(target=str(hyperlink),ref=cell_str,display=str(hyperlink_value))
        print(self._sheet[cell_str].hyperlink)
        self._sheet[cell_str].hyperlink=hyper_link
        print(self._sheet[cell_str].hyperlink)
        print(self._sheet[cell_str].value)
        
        
    def write_datetime(self,row_id,column_id,date_format,date_value,excel_style):
        datetime_value=date_value.strftime(date_format)
        self.write_cell(row_id,column_id,datetime_value,excel_style)
    
    def write_cell_with_formula(self,row_id,column_id,formula_pattern):
        formula_value='=%s'%(formula_pattern)
        self.write_cell(row_id, column_id, formula_value)
    
    def merge_cells(self,row_begin,row_end,col_begin,col_end,cell_value,excel_style):
        b_cell_str=self.get_cell_str(row_begin, col_begin)
        e_cell_str=self.get_cell_str(row_end, col_end)
        self._sheet.merge_cells('%s:%s'%(b_cell_str,e_cell_str))
        self._sheet[b_cell_str]=cell_value
        self.set_style(self._sheet[b_cell_str], excel_style)
        
    def unmerge_cells(self,row_begin,row_end,col_begin,col_end):
        b_cell_str=self.get_cell_str(row_begin, col_begin)
        e_cell_str=self.get_cell_str(row_end, col_end)
        self._sheet.unmerge_cells('%s:%s'%(b_cell_str,e_cell_str))

    def save(self):
        self._workbook.save(self._filename)
        
class ExcelReader(object):
    
    def __init__(self,filename):
        self._filename=filename
        self._workbook=openpyxl.load_workbook(self._filename,data_only=True)
        self._sheet_dict={}
        for sheet_name in self._workbook.sheetnames:
            self._sheet_dict[sheet_name]=self._workbook.get_sheet_by_name(sheet_name)
        self._sheet=self._workbook.get_sheet_by_name(self._workbook.sheetnames[0])
        self._rows=self._sheet.rows
        self._cols=self._sheet.columns
        
    def set_current_sheet(self,sheet_name):
        self._sheet=self._sheet_dict[sheet_name]
        self._rows=list(self._sheet.rows)
        self._cols=list(self._sheet.columns)
        
    def get_row_cnt(self):
        return len(self._rows)
    
    def get_col_cnt(self):
        return len(self._cols)
    
    def get_rows(self,row_id):
        return self._rows[row_id]
    
    def get_rows_value(self,row_id):
        return self._rows[row_id].value
    
    def get_cols(self,col_id):
        return self._cols(col_id)
    
    def get_cols_value(self,col_id):
        return self._cols(col_id).value
    
    def get_cell_by_id(self,row_id,col_id):
        return self._sheet.cell(row=row_id,column=col_id)
    
    def get_cell_by_str(self,cell_str):
        return self._sheet[cell_str]
    
    def get_cells_by_str(self,begin_str,end_str):
        return self._sheet[begin_str:end_str]
    
    def get_col_str_by_id(self,col_id):
        return get_column_letter(col_id)
    
    def get_col_id_by_str(self,col_str):
        return column_index_from_string(col_str)
    
    def get_hyperlink(self,row_id,col_id):
        link=self._sheet.hyperlink_map.get((row_id,col_id))
        url='(No URL)'
        if link:
            url=link.url_or_path
        return url
    
    def copy(self,new_file_name):
        wb=copy(self._workbook)
        wb.save(new_file_name)


class SimplePdfReader(object):
    
    def __init__(self,pdf_params):
        self.pdfparams=pdf_params
        self.pdffp=open(pdf_params.pdffile,'rb')
        self.pdfreader = PdfFileReader(self.pdffp)
        if self.pdfreader.isEncrypted:
            self.pdfreader.decrypt(pdf_params.password)

    def get_document_info(self):
        return self.pdfreader.getDocumentInfo()
    
    def get_page_obj(self,pageno):
        return self.pdfreader.getPage(pageno)
    
    def get_page_cnt(self):
        return self.pdfreader.numPages
    
    def get_page_text(self,pageno):
        return self.get_page_obj(pageno).extractText()
    
    def get_page_texts(self):
        content=""
        for i in range(0,self.get_page_cnt()):
            content=F"{self.get_page_text(i)}\n"
        return content
    
    def split_pdf(self,from_page):
        pdfwriter = PdfFileWriter()
        for index in range(from_page,self.get_page_cnt()):
            page_obj=self.get_page_obj(index)
            pdfwriter.addPage(page_obj)
        pdfwriter.write(open(self.pdfparams.outfile,'wb'))
    
    def close(self):
        self.pdffp.close()
            
class SimplePdfWriter(object):
    
    def __init__(self,pdf_file,font_dicts={},default_fonts={}):
        self.pdf_path=pdf_file
        self.doc=SimpleDocTemplate(self.pdf_path)
        self.font_attrs=default_fonts
        if font_dicts:
            for font_key,font_val in font_dicts.items():
                self.register_font(font_key,font_val)
        style=getSampleStyleSheet()
        self.stylesheet={}
        self.stylesheet['Title']=style['Title']
        self.stylesheet['Normal']=style['Normal']
        self.stylesheet['Heading1']=style['Heading1']
        self.stylesheet['Heading2']=style['Heading2']
        self.stylesheet['Heading3']=style['Heading3']
        self.stylesheet['Heading4']=style['Heading4']
        self.stylesheet['Heading5']=style['Heading5']
        self.stylesheet['Heading6']=style['Heading6']
        self.stylesheet['BodyText']=style['BodyText']
        self.stylesheet['BodyText'].wordWrap='CJK'
        self.stylesheet['BodyText'].firstLineIndent=32
        self.stylesheet['BodyText'].leading=30
        self.stylesheet['Table']=TableStyle(
            [('FONTNAME',(0,0),(-1,-1),'Helvetica'),
             ('FONTSIZE',(0,0),(-1,0),15),
             ('BACKGROUND',(0,0),(-1,0),HexColor('#d5dae6')),
             ('ALIGN',(0,0),(-1,-1),'CENTER'),
             ('TEXTCOLOR',(0,0),(-1,-1),colors.black),
             ('GRID',(0,0),(-1,-1),0.5,colors.grey),
             ]
            )
    
    def register_font(self,font_name,font_file):
        pdfmetrics.registerFont(TTFont(font_name,font_file))
    
    def get_table(self,pdf_item):
        dist_list=pdf_item.itemdicts['data']
        col_width=pdf_item.width
        col_height=pdf_item.height
        table_style=pdf_item.style
        table_obj=Table(dist_list,len(dist_list[0])*col_width*inch,len(dist_list)*col_height*inch)
        table_obj.setStyle(table_style)
        return table_obj
    
    def get_paragraph(self,pdf_item):
        text=pdf_item.itemdicts['text']
        text_style=pdf_item.style
        return Paragraph(text,text_style)
    
    def get_image(self,pdf_item):
        img = Image(pdf_item.item['img_path'])
        img.drawWidth=pdf_item.width*inch
        img.drawHeight=pdf_item.height*inch
        return img
    
    def get_spacer(self,pdf_item):
        return Spacer(1,pdf_item.height*inch)
    
    def get_circle(self,pdf_item):
        return Circle(pdf_item.xinch,pdf_item.yinch,pdf_item.width,fillColor=None)
    
    def get_rect(self,pdf_item):
        return Rect(pdf_item.xinch,pdf_item.yinch,pdf_item.width,pdf_item.height,fillColor=None)
    
    def get_bar(self,pdf_item):
        bc = VerticalBarChart()
        bc.x=pdf_item.xinch
        bc.y=pdf_item.yinch
        bc.height=pdf_item.height
        bc.width=pdf_item.width
        bc.data=pdf_item.itemdicts['data']
        bc.strokeColor=colors.black
        bc.valueAxis.valueMin=0
        bc.valueAxis.valueMax=pdf_item.value_max_x
        bc.valueAxis.valueStep=pdf_item.value_step
        bc.categoryAxis.labels.dx=-5
        bc.categoryAxis.labels.dy=-5
        bc.categoryAxis.categoryNames=pdf_item.itemdicts['ax_names']
        return bc
    
    def get_pie(self,pdf_item):
        pie=Pie()
        pie.x=pdf_item.xinch
        pie.y=pdf_item.yinch
        pie.slices.label_boxStrokeColor=colors.white
        
        pie.data=pdf_item.itemdicts['data']
        pie.labels=pdf_item.itemdicts['labels']
        pie.simpleLabels=0
        pie.sameRadii=1
        
        pie.slices._strokeColor=colors.blue
        pie.strokeWidth=1
        pie.strokeColor=colors.white
        pie.slices.label_pointer_piePad=10
        pie.slices.label_pointer_edgePad=25
        pie.width = pdf_item.width
        pie.direction='clockwise'
        pie.pointerLabelMode='LeftRight'
        if 'font_name' in pdf_item.itemdicts.keys():
            for i in range(len(pdf_item.itemdicts['labels'])):
                pie.slices[i].fontName=pdf_item.itemdicts['font_name']
        for i , col in enumerate(pdf_item.itemdicts['colors']):
            pie.slices[i].fillColor=col
        return pie
    
    def get_lineplot(self,pdf_item):
        lp=LinePlot()
        lp.x=pdf_item.xinch
        lp.y=pdf_item.yinch
        lp.height=pdf_item.height
        lp.width=pdf_item.width
        lp.joinedLines=1
        lp.data=pdf_item.itemdicts['data']
        lp.lineLabelFormat='%2.0f'
        lp.strokeColor=colors.black
        for i, line_items in enumerate(pdf_item.itemdicts['attrs']):
            lp.lines[i].strokeColor=line_items[0]
            lp.lines[i].symbol = makeMarker(line_items[1])
        lp.xValueAxis.valueMin=0
        lp.xValueAxis.valueMax=pdf_item.value_max_x
        lp.xValueAxis.valueStep=pdf_item.value_step
        lp.yValueAxis.valueMin=0
        lp.yValueAxis.valueMax=pdf_item.value_max_y
        lp.yValueAxis.valueStep=pdf_item.value_step
        return lp
        
    
    def get_legend(self,pdf_item):
        leg=Legend()
        leg.fontName='Helvetica'
        leg.alignment='right'
        leg.boxAnchor='ne'
        leg.x=pdf_item.xinch
        leg.y=pdf_item.yinch
        leg.dxTextSpace=10
        leg.columnMaximum=2
        leg.colorNamePairs=pdf_item.itemdicts['leg_items']
        return leg
    
    def get_label(self,pdf_item):
        lab=Label()
        lab.x=pdf_item.xinch
        lab.y=pdf_item.yinch
        title=pdf_item.itemdicts['title']
        lab.setText(title)
        if 'font_name' in pdf_item.itemdicts.keys():
            lab.fontName=pdf_item.itemdicts['font_name']
        lab.fontSize=20
        return lab
    
    def get_drawing_with_complex_charts(self,pdf_items):
        drawing=self.get_drawing(pdf_items[0])
        lab = self.get_label(pdf_items[0])
        item_type=pdf_items[1].itemtype
        if item_type not in ['bar','pie','lineplot']:
            drawing.add(lab)
            return drawing
        else:
            func=F"get_{item_type}"
            chart=self.func(pdf_items[1])
            leg=self.get_legend(pdf_items[2])
            drawing.add(lab)
            drawing.background=Rect(0,0,pdf_items[0].width.pdf_items[0].height,strokeWidth=1,strokeColor="#868686",fillColor=None)
            drawing.add(leg)
            drawing.add(chart)
            return drawing
        
    def get_drawing(self,pdf_item):
        return Drawing(pdf_item.width,pdf_item.height)
    
    def write_pages(self,stories):
        page_brk=PageBreak()
        contents=[]
        for storys in stories:
            for story in storys:
                contents.append(story)
            contents.append(page_brk)
        self.doc.build(contents)
        return
#     _FilledCircle = _doFill
#     _FilledSquare = _doFill
#     _FilledDiamond = _doFill
#     _FilledCross = _doFill
#     _FilledTriangle = _doFill
#     _FilledStarSix = _doFill
#     _FilledPentagon = _doFill
#     _FilledHexagon = _doFill
#     _FilledHeptagon = _doFill
#     _FilledOctagon = _doFill
#     _FilledStarFive = _doFill
#     _FilledArrowHead = _doFill    
def merge_pdf(in_list, out_file):
    pdf_writer = PdfFileWriter()
    for in_file in in_list:
        pdf_reader = PdfFileReader(open(in_file, 'rb'))
        num_pages = pdf_reader.getNumPages()
        for index in range(0, num_pages):
            page_obj = pdf_reader.getPage(index)
            pdf_writer.addPage(page_obj)
        pdf_writer.write(open(out_file, 'wb'))

if __name__=='__main__':
    pass
#    sheet_names=['test1','sheet1','value1']
#    excel_writer=ExcelWriter('C:/Users/xcKev/eclipse-workspace/KToolApps/test/test.xlsx',sheet_names=sheet_names)
#    excel_writer.set_current_sheet('value1')
#    excel_style=ExcelStyle()
#    rows1=['hah','value','mytest']
#    rows2=[1,2,4,5,39,20]
#    columns1=[1230,324432,232432,345]
#    columns2=['hva','cool','clearly']
#    excel_writer.write_row(1, rows1, excel_style.get_style())
#    excel_writer.write_row(2, rows2, excel_style.get_style())
#    excel_writer.write_column(7, columns1, excel_style.get_style())
#    excel_writer.write_column(9, columns2, excel_style.get_style())
#    excel_writer.write_cell_with_formula(2, 8, 'F2+G2')
#    excel_writer.write_hyperlink(3, 8, 'http://www.baidu.com', 'baidu')
#    excel_writer.save()
#     sheet_names=['test1','sheet1','value1']
#     excel_reader=ExcelReader('C:/Users/xcKev/eclipse-workspace/KToolApps/test/test.xls')
#     excel_reader.set_current_sheet('value1')
#     print(excel_reader.get_row_cnt())
#     print(excel_reader.get_col_cnt())
#     print(excel_reader.get_cols(8))
#     print(excel_reader.get_rows(0))
#     print(excel_reader.get_cell(5,7))
#     print(excel_reader.get_hyperlink(5, 7))
#     print(excel_reader.get_formulas())
#     csv_reader=CsvReader('C:/Users/xcKev/eclipse-workspace/KToolApps/test/test.csv',_delimiter='|')
#     print(csv_reader._lines)
#     headers=csv_reader.get_headers(0)
#     print(headers)
#     print(csv_reader._total_cnt)
#     details=csv_reader.get_details(1,4)
#     print(details)
