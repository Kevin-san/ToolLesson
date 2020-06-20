# -*- encoding:UTF-8 -*-
'''
Created on 2020/5/3

@author: xcKev
'''
import time
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

class ExcelWriter(object):
     
    def __init__(self, filename, sheet_names):
        self._filename = filename
        self._workbook = openpyxl.Workbook()
        self._sheet_dict = {}
        for i, sheet_name in enumerate(sheet_names):
            sheet = self._workbook.create_sheet(sheet_name, i)
            self._sheet_dict[sheet_name] = sheet
        self._sheet = self._sheet_dict[sheet_names[0]]
    
    def set_current_sheet(self, sheet_name):
        self._sheet = self.get_sheet(sheet_name)
        
    def set_style(self, cell, excel_style):
        cell.font = excel_style.font
        cell.border = excel_style.border
        cell.alignment = excel_style.alignment
        cell.fill = excel_style.fill
        cell.protection = excel_style.protection
        
    def get_sheet(self, sheet_name):
        return self._sheet_dict[sheet_name]
    
    def remove_sheet(self, sheet_name):
        return self._workbook.remove_sheet(self._sheet_dict[sheet_name])
    
    def get_cell_str(self, row_id, col_id):
        col_str = get_column_letter(col_id)
        cell_str = col_str + str(row_id)
        return cell_str
    
    def write_row(self, row_id, rows, excel_style):
        for i in range(1, len(rows) + 1):
            cell_str = self.get_cell_str(row_id, i)
            self._sheet[cell_str] = rows[i - 1]
            self.set_style(self._sheet[cell_str], excel_style)

    def write_column(self, column_id, columns, excel_style):
        for i in range(1, len(columns) + 1):
            cell_str = self.get_cell_str(i, column_id)
            self._sheet[cell_str] = columns[i - 1]
            self.set_style(self._sheet[cell_str], excel_style)

    def append_rows(self, rows):
        self._sheet.append(rows)
    
    def append_cols(self, cols):
        self._sheet.append(list(zip(cols)))
    
    def write_cell(self, row_id, column_id, cell_value):
        cell_str = self.get_cell_str(row_id, column_id)
        self._sheet[cell_str] = cell_value
    
    def write_hyperlink(self, row_id, column_id, hyperlink, hyperlink_value):
        cell_str = self.get_cell_str(row_id, column_id)
        hyper_link = Hyperlink(target=str(hyperlink), ref=cell_str, display=str(hyperlink_value))
        print(self._sheet[cell_str].hyperlink)
        self._sheet[cell_str].hyperlink = hyper_link
        print(self._sheet[cell_str].hyperlink)
        print(self._sheet[cell_str].value)
        
        
    def write_datetime(self, row_id, column_id, date_format, date_value, excel_style):
        datetime_value = time.strftime(date_format,date_value)
        self.write_cell(row_id, column_id, datetime_value)
        cell_str=self.get_cell_str(row_id, column_id)
        self.set_style(self._sheet[cell_str], excel_style)
    
    def write_cell_with_formula(self, row_id, column_id, formula_pattern):
        formula_value = '=%s' % (formula_pattern)
        self.write_cell(row_id, column_id, formula_value)
    
    def merge_cells(self, row_begin, row_end, col_begin, col_end, cell_value, excel_style):
        b_cell_str = self.get_cell_str(row_begin, col_begin)
        e_cell_str = self.get_cell_str(row_end, col_end)
        print('%s:%s'%(b_cell_str,e_cell_str))
        self._sheet.merge_cells('%s:%s' % (b_cell_str, e_cell_str))
        self._sheet[b_cell_str] = cell_value
        self.set_style(self._sheet[b_cell_str], excel_style)
        
    def unmerge_cells(self, row_begin, row_end, col_begin, col_end):
        b_cell_str = self.get_cell_str(row_begin, col_begin)
        e_cell_str = self.get_cell_str(row_end, col_end)
        print('%s:%s'%(b_cell_str,e_cell_str))
        self._sheet.unmerge_cells('%s:%s' % (b_cell_str, e_cell_str))

    def save(self):
        self._workbook.save(self._filename)
